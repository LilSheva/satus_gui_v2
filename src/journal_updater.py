# ==================================================================================
# МОДУЛЬ 7: ОБНОВЛЕНИЕ ЖУРНАЛА (Версия 5, с исправлением потери статуса "Условно")
# ==================================================================================
import pandas as pd
from datetime import datetime, timedelta
import os
import shutil
import openpyxl
from openpyxl.styles import Font


def generate_new_journal_name(original_path: str) -> str:
    # ... (код без изменений) ...
    now = datetime.now()
    base_name = "Журнал публикаций уязвимостей"
    file_ext = ".xlsx"
    if 8 <= now.hour < 20:
        date_str = now.strftime("%d.%m.%Y");
        suffix = ""
    elif 20 <= now.hour <= 23:
        date_str = now.strftime("%d.%m.%Y");
        suffix = " (2)"
    else:
        yesterday = now - timedelta(days=1);
        date_str = yesterday.strftime("%d.%m.%Y");
        suffix = " (2)"
    dir_name = os.path.dirname(original_path)
    new_name = f"{base_name} {date_str}{suffix}{file_ext}"
    return os.path.join(dir_name, new_name)


def _find_first_data_row(sheet: any) -> int:
    # ... (код без изменений) ...
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if isinstance(cell_value, (int, float)):
            return row
    return 2


def update_journal_file(original_journal_path: str, verified_report_path: str) -> pd.DataFrame:
    try:
        # 1. Загрузка данных
        verified_df = pd.read_excel(verified_report_path, sheet_name='Основная таблица', dtype=str)
        verified_df.dropna(subset=['CVE'], inplace=True)  # Удаляем строки совсем без CVE
        # Заменяем возможные NaN (пустые ячейки) на пустые строки для дальнейшей обработки
        verified_df.fillna('', inplace=True)

        # 2. Очистка, унификация и фильтрация
        verified_df['Статус'] = verified_df[
            'Статус'].str.strip().str.upper()  # Убираем пробелы и приводим к ВЕРХНЕМУ РЕГИСТРУ
        # Отфильтровываем строки, где статус пустой
        verified_df = verified_df[verified_df['Статус'] != ''].copy()
        if verified_df.empty:
            print("Нет данных для добавления в Журнал (все статусы пустые).")
            return pd.DataFrame()

        # 3. Сортировка на унифицированных данных
        # Теперь все статусы в верхнем регистре, как и в списке
        status_order = ["ДА", "УСЛОВНО", "LINUX", "НЕТ", "ПОВТОР"]
        category_type = pd.CategoricalDtype(categories=status_order, ordered=True)
        verified_df['Статус'] = verified_df['Статус'].astype(category_type)

        # Удаляем строки, которые стали NaN (если в Excel был статус, которого нет в списке, например, "Возможно")
        verified_df.dropna(subset=['Статус'], inplace=True)

        verified_df.sort_values('Статус', inplace=True)

        # 4. Анализ структуры старого ЖП
        original_workbook = openpyxl.load_workbook(original_journal_path)
        original_sheet = original_workbook.active
        first_data_row = _find_first_data_row(original_sheet)
        header_rows_to_skip = first_data_row - 1
        old_journal_df = pd.read_excel(original_journal_path, sheet_name=0, skiprows=header_rows_to_skip)

        # 5. Нумерация снизу вверх
        last_num = pd.to_numeric(old_journal_df.iloc[:, 0], errors='coerce').max()
        if pd.isna(last_num): last_num = 0
        n = len(verified_df)
        new_nums = range(int(last_num) + 2, int(last_num) + n + 2)
        verified_df['№'] = sorted(new_nums, reverse=True)

        # 6. Создание нового файла
        new_journal_path = generate_new_journal_name(original_journal_path)
        shutil.copy(original_journal_path, new_journal_path)

        # 7. "Внедрение" строк в новую копию
        workbook = openpyxl.load_workbook(new_journal_path)
        sheet = workbook.active
        sheet.insert_rows(idx=first_data_row, amount=n + 1)

        cols_order = ['№', 'Дата обработки', 'Ответственный', 'Публикация', 'Статус', 'ID ППТС', 'CVE', 'CVSS',
                      'Продукт', 'Источник']
        data_to_insert = verified_df[cols_order].values.tolist()

        for row_idx_offset, row_data in enumerate(data_to_insert):
            current_row = first_data_row + row_idx_offset
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=current_row, column=col_idx, value=cell_value)

        # 8. Форматирование статусов
        status_colors = {"ДА": "FF0000", "ПОВТОР": "FF0000", "УСЛОВНО": "FFA500", "LINUX": "0070C0", "НЕТ": "008000"}
        status_col_letter = 'E'
        for row in range(first_data_row, first_data_row + n):
            cell = sheet[f'{status_col_letter}{row}']
            if cell.value in status_colors:
                cell.font = Font(color=status_colors[cell.value])

        workbook.save(new_journal_path)

        print(f"Журнал обновлен и сохранен как: {new_journal_path}")
        return verified_df

    except FileNotFoundError as e:
        print(f"ОШИБКА: Файл не найден: {e.filename}")
    except Exception as e:
        print(f"ОШИБКА при обновлении Журнала Публикаций: {e}")
    return pd.DataFrame()


if __name__ == '__main__':
    VERIFIED_REPORT_PATH = "C:/Users/yakov/PycharmProjects/test4/status_v2/res_tmp_test_report.xlsx"
    JOURNAL_PATH = "C:/Users/yakov/PycharmProjects/test4/status_v2/Журнал публикаций уязвимостей 15.10.2025.xlsx"

    print("--- Тестирование модуля journal_updater (v5, 'пуленепробиваемая' версия) ---")
    if os.path.exists(VERIFIED_REPORT_PATH) and os.path.exists(JOURNAL_PATH):
        added_data_df = update_journal_file(JOURNAL_PATH, VERIFIED_REPORT_PATH)
        if not added_data_df.empty:
            print("\nДанные, которые были добавлены (уже отсортированы):")
            print(added_data_df[['№', 'Статус', 'CVE']].to_string())
    else:
        print("\nОШИБКА: Не найден один из тестовых файлов для запуска.")